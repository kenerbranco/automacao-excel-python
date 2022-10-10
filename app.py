import openpyxl

# Criar um planilha(book)
book = openpyxl.Workbook()
# Como visualizar páginas existentes
print(book.sheetnames)
# Como criar uma página
book.create_sheet('Frutas')
# Como selecionar um página
frutas_page = book['Frutas']
frutas_page.append(['Frutas', 'Qtd', 'Preç'])
frutas_page.append(['Banana', '5', 'R$3,90'])
frutas_page.append(['Fruta 2', '2', 'R$15,90'])
frutas_page.append(['Fruta 3', '10', 'R$30,90'])
frutas_page.append(['Fruta 4', '2', 'R$50,50'])
book.save('Planilha de Compras.xlsx')
