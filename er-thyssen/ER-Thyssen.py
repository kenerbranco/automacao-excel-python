"""
    Nome ponto - Nome tab. final

    1Hv+ -> 1H - motor radial
    1He3 -> 1H - motor radial (env)
    2Av -> 2A - motor axial
    2Ae3 -> 2A - motor axial (env)
    2Hv -> 2H - motor radial
    2He3 -> 2H - motor radial (env)
    3Av -> 3A - redutor axial
    3Ae3 - > 3A - redutor axial (env)
    3Hv -> 3H - redutor radial
    3He3 -> 3H - redutor radial (env)
    4Av -> 4A - redutor axial
    4Ae3 -> 4A - redutor axial (env)
    4Hv -> 4H - redutor radial
    4He3 -> 4H - redutor radial (env)
    5Av -> 5A - redutor axial
    5Ae3 -> 5A - redutor axial (env)
    5Hv -> 5H - redutor radial
    5He3 -> 5H - redutor radial (env)
    6Hv -> 6H - redutor radial
    6He3 -> 6H - redutor radial (env)
"""

import openpyxl
from datetime import date
from time import sleep

maq = dict()

# ER01
maq_final = [   
                {'Ponto' : [0,0,0,0,0,'Alteração','Unidade','Última medição']}, 
                {'1H - motor radial' : [0,0,0,0,0,0,0,0]}, 
                {'1H - motor radial (env)' : [0,0,0,0,0,0,0,0]}, 
                {'2A - motor axial' : [0,0,0,0,0,0,0,0]}, 
                {'2A - motor axial (env)' : [0,0,0,0,0,0,0,0]}, 
                {'2H - motor radial' : [0,0,0,0,0,0,0,0]}, 
                {'2H - motor radial (env)' : [0,0,0,0,0,0,0,0]},
                {'3A - redutor axial' : [0,0,0,0,0,0,0,0]}, 
                {'3A - redutor axial (env)' : [0,0,0,0,0,0,0,0]}, 
                {'3H - redutor radial' : [0,0,0,0,0,0,0,0]}, 
                {'3H - redutor radial (env)' : [0,0,0,0,0,0,0,0]},
                {'4A - redutor axial' : [0,0,0,0,0,0,0,0]}, 
                {'4A - redutor axial (env)' : [0,0,0,0,0,0,0,0]}, 
                {'4H - redutor radial' : [0,0,0,0,0,0,0,0]}, 
                {'4H - redutor radial (env)' : [0,0,0,0,0,0,0,0]}, 
                {'5A - redutor axial' : [0,0,0,0,0,0,0,0]}, 
                {'5A - redutor axial (env)' : [0,0,0,0,0,0,0,0]},
                {'5H - redutor radial' : [0,0,0,0,0,0,0,0]}, 
                {'5H - redutor radial (env)' : [0,0,0,0,0,0,0,0]}, 
                {'6H - redutor radial' : [0,0,0,0,0,0,0,0]}, 
                {'6H - redutor radial (env)' : [0,0,0,0,0,0,0,0]}
            ]

# ER02
maq_final2 = [  
                {'Ponto' : [0,0,0,0,0,'Alteração','Unidade','Última medição']}, 
                {'1H - motor radial' : [0,0,0,0,0,0,0,0]}, 
                {'1H - motor radial (env)' : [0,0,0,0,0,0,0,0]}, 
                {'2A - motor axial' : [0,0,0,0,0,0,0,0]}, 
                {'2A - motor axial (env)' : [0,0,0,0,0,0,0,0]}, 
                {'2H - motor radial' : [0,0,0,0,0,0,0,0]}, 
                {'2H - motor radial (env)' : [0,0,0,0,0,0,0,0]},
                {'3A - redutor axial' : [0,0,0,0,0,0,0,0]}, 
                {'3A - redutor axial (env)' : [0,0,0,0,0,0,0,0]}, 
                {'3H - redutor radial' : [0,0,0,0,0,0,0,0]}, 
                {'3H - redutor radial (env)' : [0,0,0,0,0,0,0,0]},
                {'4A - redutor axial' : [0,0,0,0,0,0,0,0]}, 
                {'4A - redutor axial (env)' : [0,0,0,0,0,0,0,0]}, 
                {'4H - redutor radial' : [0,0,0,0,0,0,0,0]}, 
                {'4H - redutor radial (env)' : [0,0,0,0,0,0,0,0]}, 
                {'5A - redutor axial' : [0,0,0,0,0,0,0,0]}, 
                {'5A - redutor axial (env)' : [0,0,0,0,0,0,0,0]},
                {'5H - redutor radial' : [0,0,0,0,0,0,0,0]}, 
                {'5H - redutor radial (env)' : [0,0,0,0,0,0,0,0]}, 
                {'6H - redutor radial' : [0,0,0,0,0,0,0,0]}, 
                {'6H - redutor radial (env)' : [0,0,0,0,0,0,0,0]}
            ]

# Leitura arquivo .xlsx
book = openpyxl.load_workbook("index.xlsx")

# Atribuindo sheets em variáveis
index = book['index']
book.create_sheet('tabela-final-motor01')
book.create_sheet('tabela-final-motor02')

# maximo de linhas e colunas
max_lin = index.max_row
max_col = index.max_column

def extCabecalho(l):  # Armazena dados cabeçalho
    lst = list()
    for la in range(l, l+5):
        cel = index.cell(row=la, column=15)
        val = cel.value
        lst.append(val)
    lst.append('Unidade')
    lst.append('Alteração')
    lst.append('Última medição')
    return lst


def extDados(l):
    lst = list()
    for la in range(l, l+5):  # Armazena cinco medições do ponto
        cel = index.cell(row=la, column=7)
        val = cel.value
        lst.append(val)
    cel_unidade = index.cell(row=l, column=9)  # Unidade medida do ponto
    cel_dif = index.cell(row=l, column=11)  # Diferença em porcentagem da última medição
    cel_ultMedicao = index.cell(row=l, column=5)  # Data última medição
    lst.append(cel_unidade.value)
    lst.append(cel_dif.value)
    lst.append(cel_ultMedicao.value)
    return lst


def nomeArq(l):  # Armazena nome do equipamento para salvar planilha posteriormente
    cel_nome = index.cell(row=l, column=1)
    valor = cel_nome.value.split()
    cel_ultMedicao = index.cell(row=l, column=5)
    data = str(cel_ultMedicao.value).replace('/', '-')
    nome = valor[1] + '-' + valor[2] + '-' + valor[3] + '-' + data[:10]
    return nome


def nomeArqErr(): # Armazena nome padrão caso dê erro "def nomeArq()"
    data = date.today()
    nome = str(data)
    return nome


# Programa principal
print('-' * 42)
print('{:^42}'.format(' Automação Planilha v1.0 '))
print('{:^42}'.format(' ER Thyssen '))
print('-' * 42)

sleep(0.8)
try: # Novo valor de max_lin para range de apenas dados necessários
    for l in range(1, max_lin+1): 
        for c in range(1, max_col):
            celula = index.cell(row=l, column=c)
            valor = str(celula.value).split()
            if 'Notas' in valor:
                max_lin = l
except:
    print('  => [ ERRO ] - Filtro linhas necessárias')
else:
    print('  => [ OK ] - Filtro linhas necessárias')

sleep(0.8)
try: # Extração dos pontos
    for l in range(1, max_lin+1):
        for c in range(1, max_col):
            maq.clear()

            # Célula motor/redutor atual
            celulaMotor = index.cell(row=l, column=1)
            valorMotor = str(celulaMotor.value).split()

            # Célula nome do ponto
            celula = index.cell(row=l, column=c)
            valor = str(celula.value).split()

            # Extração dados Motor 01
            if ('1Hv+' in valor) and ('MOT1' or 'MOT' in valorMotor):
                maq['Ponto'] = extCabecalho(l)
                maq_final[0] = maq.copy()
                maq.clear()
                maq['1H - motor radial'] = extDados(l)
                maq_final[1] = maq.copy()
                nome_arq = nomeArq(l) # Captura nome do arquivo
            if ('1He3' in valor) and ('MOT1' or 'MOT' in valorMotor):
                maq['1H - motor radial (env)'] = extDados(l)
                maq_final[2] = maq.copy()
            if ('2Av' in valor) and ('MOT1' or 'MOT' in valorMotor):
                maq['2A - motor axial'] = extDados(l)
                maq_final[3] = maq.copy()
            if ('2Ae3' in valor) and ('MOT1' or 'MOT' in valorMotor):
                maq['2A - motor axial (env)'] = extDados(l)
                maq_final[4] = maq.copy()
            if ('2Hv' in valor) and ('MOT1' or 'MOT' in valorMotor):
                maq['2H - motor radial'] = extDados(l)
                maq_final[5] = maq.copy()
            if ('2He3' in valor) and ('MOT1' or 'MOT' in valorMotor):
                maq['2H - motor radial (env)'] = extDados(l)
                maq_final[6] = maq.copy()
            if ('3Av' in valor) and ('RED1' or 'RED' in valorMotor):
                maq['3A - redutor axial'] = extDados(l)
                maq_final[7] = maq.copy()
            if ('3Ae3' in valor) and ('RED1' or 'RED' in valorMotor):
                maq['3A - redutor axial (env)'] = extDados(l)
                maq_final[8] = maq.copy()
            if ('3Hv' in valor) and ('RED1' or 'RED' in valorMotor):
                maq['3H - redutor radial'] = extDados(l)
                maq_final[9] = maq.copy()
            if ('3He3' in valor) and ('RED1' or 'RED' in valorMotor):
                maq['3H - redutor radial (env)'] = extDados(l)
                maq_final[10] = maq.copy()
            if ('4Av' in valor) and ('RED1' or 'RED' in valorMotor):
                maq['4A - redutor axial'] = extDados(l)
                maq_final[11] = maq.copy()
            if ('4Ae3' in valor) and ('RED1' or 'RED' in valorMotor):
                maq['4A - redutor axial (env)'] = extDados(l)
                maq_final[12] = maq.copy()
            if ('4Hv' in valor) and ('RED1' or 'RED' in valorMotor):
                maq['4H - redutor radial'] = extDados(l)
                maq_final[13] = maq.copy()
            if ('4He3' in valor) and ('RED1' or 'RED' in valorMotor):
                maq['4H - redutor radial (env)'] = extDados(l)
                maq_final[14] = maq.copy()
            if ('5Av' in valor) and ('RED1' or 'RED' in valorMotor):
                maq['5A - redutor axial'] = extDados(l)
                maq_final[15] = maq.copy()
            if ('5Ae3' in valor) and ('RED1' or 'RED' in valorMotor):
                maq['5A - redutor axial (env)'] = extDados(l)
                maq_final[16] = maq.copy()
            if ('5Hv' in valor) and ('RED1' or 'RED' in valorMotor):
                maq['5H - redutor radial'] = extDados(l)
                maq_final[17] = maq.copy()
            if ('5He3' in valor) and ('RED1' or 'RED' in valorMotor):
                maq['5H - redutor radial (env)'] = extDados(l)
                maq_final[18] = maq.copy()
            if ('6Hv' in valor) and ('RED1' or 'RED' in valorMotor):
                maq['6H - redutor radial'] = extDados(l)
                maq_final[19] = maq.copy()
            if ('6He3' in valor) and ('RED1' or 'RED' in valorMotor):
                maq['6H - redutor radial (env)'] = extDados(l)
                maq_final[20] = maq.copy()

            # Extração dados Motor 02
            if ('1Hv+' in valor) and ('MOT2' in valorMotor):
                maq['1H - motor radial'] = extDados(l)
                maq_final2[1] = maq.copy()
                maq.clear()
                maq['Ponto'] = extCabecalho(l)
                maq_final2[0] = maq.copy()
            if ('1He3' in valor) and ('MOT2' in valorMotor):
                maq['1H - motor radial (env)'] = extDados(l)
                maq_final2[2] = maq.copy()
            if ('2Av' in valor) and ('MOT2' in valorMotor):
                maq['2A - motor axial'] = extDados(l)
                maq_final2[3] = maq.copy()
            if ('2Ae3' in valor) and ('MOT2' in valorMotor):
                maq['2A - motor axial (env)'] = extDados(l)
                maq_final2[4] = maq.copy()
            if ('2Hv' in valor) and ('MOT2' in valorMotor):
                maq['2H - motor radial'] = extDados(l)
                maq_final2[5] = maq.copy()
            if ('2He3' in valor) and ('MOT2' in valorMotor):
                maq['2H - motor radial (env)'] = extDados(l)
                maq_final2[6] = maq.copy()
            if ('3Av' in valor) and ('RED2' in valorMotor):
                maq['3A - redutor axial'] = extDados(l)
                maq_final2[7] = maq.copy()
            if ('3Ae3' in valor) and ('RED2' in valorMotor):
                maq['3A - redutor axial (env)'] = extDados(l)
                maq_final2[8] = maq.copy()
            if ('3Hv' in valor) and ('RED2' in valorMotor):
                maq['3H - redutor radial'] = extDados(l)
                maq_final2[9] = maq.copy()
            if ('3He3' in valor) and ('RED2' in valorMotor):
                maq['3H - redutor radial (env)'] = extDados(l)
                maq_final2[10] = maq.copy()
            if ('4Av' in valor) and ('RED2' in valorMotor):
                maq['4A - redutor axial'] = extDados(l)
                maq_final2[11] = maq.copy()
            if ('4Ae3' in valor) and ('RED2' in valorMotor):
                maq['4A - redutor axial (env)'] = extDados(l)
                maq_final2[12] = maq.copy()
            if ('4Hv' in valor) and ('RED2' in valorMotor):
                maq['4H - redutor radial'] = extDados(l)
                maq_final2[13] = maq.copy()
            if ('4He3' in valor) and ('RED2' in valorMotor):
                maq['4H - redutor radial (env)'] = extDados(l)
                maq_final2[14] = maq.copy()
            if ('5Av' in valor) and ('RED2' in valorMotor):
                maq['5A - redutor axial'] = extDados(l)
                maq_final2[15] = maq.copy()
            if ('5Ae3' in valor) and ('RED2' in valorMotor):
                maq['5A - redutor axial (env)'] = extDados(l)
                maq_final2[16] = maq.copy()
            if ('5Hv' in valor) and ('RED2' in valorMotor):
                maq['5H - redutor radial'] = extDados(l)
                maq_final2[17] = maq.copy()
            if ('5He3' in valor) and ('RED2' in valorMotor):
                maq['5H - redutor radial (env)'] = extDados(l)
                maq_final2[18] = maq.copy()
            if ('6Hv' in valor) and ('RED2' in valorMotor):
                maq['6H - redutor radial'] = extDados(l)
                maq_final2[19] = maq.copy()
            if ('6He3' in valor) and ('RED2' in valorMotor):
                maq['6H - redutor radial (env)'] = extDados(l)
                maq_final2[20] = maq.copy()
except:
    print('  => [ ERRO ] - Extração dos pontos')
else:
    print('  => [ OK ] - Extração dos pontos')

sleep(0.8)
try: # Gravação dados novo sheet referente motor01
    final = book['tabela-final-motor01']
    lin_final = col_final = 1
    for ponto in maq_final:
        for k, v in ponto.items():
            final.cell(row=lin_final, column=col_final).value = k
            col_final += 1
            for medida in v:
                final.cell(row=lin_final, column=col_final).value = medida
                col_final += 1
        lin_final += 1
        col_final = 1    
except:
    print('  => [ ERRO ] - Criação tabela Motor 01')
else:
    print('  => [ OK ] - Criação tabela Motor 01')

sleep(0.8)
try: # Gravação dados novo sheet referente motor02
    final = book['tabela-final-motor02']
    lin_final = col_final = 1
    for ponto in maq_final2:
        for k, v in ponto.items():
            final.cell(row=lin_final, column=col_final).value = k
            col_final += 1
            for medida in v:
                final.cell(row=lin_final, column=col_final).value = medida
                col_final += 1
        lin_final += 1
        col_final = 1    
except:
    print('  => [ ERRO ] - Criação tabela Motor 02')
else:
    print('  => [ OK ] - Criação tabela Motor 02')

sleep(0.8)
try: # Salvar arquivo formato xlsx
    book.save(f'{nome_arq}.xlsx')
except:
    dataName = nomeArqErr()
    book.save(f'{dataName}.xlsx')
    print('  => [ ERRO ] - Criação novo arquivo nome estação".xlsx"')
    print('     => [ OK ] - Criação novo arquivo ".xlsx" c/ data atual')
else:
    print('  => [ OK ] - Criação novo arquivo ".xlsx"')

print('-' * 42)
input('<< Pressione qualquer tecla para sair >>')
