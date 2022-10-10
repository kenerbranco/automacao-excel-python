import openpyxl

# Carregando arquivo
book = openpyxl.load_workbook('Planilha de Compras.xlsx')
# Selecionando um página
frutas_page = book['Frutas']
# Imprimindo os dados de cada linha
for rows in frutas_page.iter_rows(min_row=2, max_row=5):
    print(rows[0].value, rows[1].value, rows[2].value)
#    for cell in rows:
#        print(cell.value)

for rows in frutas_page.iter_rows(min_row=2, max_row=5):
    for cell in rows:
        if cell.value == 'Banana':
            cell.value = 'Fruta 1'

# Salvar as alterações
book.save('Planilha de Compras v2.xlsx')
