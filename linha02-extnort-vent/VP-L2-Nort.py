"""
    Nome ponto - Nome tab. final

    1Ae3+ - 1A-Motor Axial (env.)
    1Av - 1A-Motor Axial
    1He3 - 1H-Motor Radial (env.)
    1Hv+ - 1H-Motor Radial
    2Ve3+ - 2H-Motor Radial (env.)
    2Vv - 2H-Motor Radial
    Axial - Carc.Axial-Central
    Radial - Carc.Radial-Central
    BASE01 - Base-01
    BASE02 - Base-02
    BASE03 - Base-03
    BASE04 - Base-04
"""

import openpyxl
from datetime import date
from time import sleep

maq = dict()
maq_final = [   
                {'Ponto' : [0,0,0,0,0,'Alteração','Unidade','Última medição']}, 
                {'Motor/Axial (envelope)' : [0,0,0,0,0,0,0,0]}, 
                {'Motor/Axial' : [0,0,0,0,0,0,0,0]}, 
                {'Motor/Radial (envelope)' : [0,0,0,0,0,0,0,0]}, 
                {'Motor/Radial' : [0,0,0,0,0,0,0,0]}, 
                {'Carcaça/Radial' : [0,0,0,0,0,0,0,0]}, 
                {'Carcaça/Axial' : [0,0,0,0,0,0,0,0]},
                {'Lewa/Atuem/Axial(envelope)' : [0,0,0,0,0,0,0,0]}, 
                {'Lewa/Atuem/Axial' : [0,0,0,0,0,0,0,0]}, 
                {'Lewa/Atuem/Radial(envelope)' : [0,0,0,0,0,0,0,0]},
                {'Base-01' : [0,0,0,0,0,0,0,0]}, 
                {'Base-02' : [0,0,0,0,0,0,0,0]}, 
                {'Base-03' : [0,0,0,0,0,0,0,0]}, 
                {'Base-04' : [0,0,0,0,0,0,0,0]}
            ]

# Leitura arquivo .xlsx
book = openpyxl.load_workbook("index.xlsx")

# Atribuindo sheets em variáveis
index = book['index']
book.create_sheet('tabela-final')

# maximo de linhas e colunas
max_lin = index.max_row
max_col = index.max_column

def extCabecalho(l):  # Armazena dados cabeçalho
    lst = list()
    for la in range(l, l+5):
        cel = index.cell(row=la, column=15)
        val = cel.value
        lst.append(val)
    lst.append('Unidade')
    lst.append('Alteração')
    lst.append('Última medição')
    return lst


def extDados(l):
    lst = list()
    for la in range(l, l+5):  # Armazena cinco medições do ponto
        cel = index.cell(row=la, column=7)
        val = cel.value
        lst.append(val)
    cel_unidade = index.cell(row=l, column=9)  # Unidade medida do ponto
    cel_dif = index.cell(row=l, column=11)  # Diferença em porcentagem da última medição
    cel_ultMedicao = index.cell(row=l, column=5)  # Data última medição
    lst.append(cel_unidade.value)
    lst.append(cel_dif.value)
    lst.append(cel_ultMedicao.value)
    return lst


def nomeArq(l):  # Armeza nome do equipamento para salvar planilha posteriormente
    cel_nome = index.cell(row=l, column=1)
    valor = cel_nome.value.split()
    cel_ultMedicao = index.cell(row=l, column=5)
    data = str(cel_ultMedicao.value).replace('/', '-')
    nome = valor[0] + '-' + valor[2] + '-' + data[:10]
    return nome


def nomeArqErr(): # Armazena nome padrão caso dê erro "def nomeArq()"
    data = date.today()
    nome = str(data)
    return nome


# Programa principal
print('-' * 42)
print('{:^42}'.format(' Automação Planilha v1.0 '))
print('{:^42}'.format(' Ventilação L2 - L1 norte '))
print('-' * 42)

sleep(0.8)
try: # Novo valor de max_lin para range de apenas dados necessários
    for l in range(1, max_lin+1): 
        for c in range(1, max_col):
            celula = index.cell(row=l, column=c)
            valor = str(celula.value).split()
            if 'Notas' in valor:
                max_lin = l
except:
    print('  => [ ERRO ] - Filtro linhas necessárias')
else:
    print('  => [ OK ] - Filtro linhas necessárias')

sleep(0.8)
try: # Extração dos pontos
    for l in range(1, max_lin+1):
        for c in range(1, max_col):
            maq.clear()
            celula = index.cell(row=l, column=c)
            valor = str(celula.value).split()
            if '1Ae3+' in valor:
                maq['1A-Motor Axial (env.)'] = extDados(l)
                maq_final[1] = maq.copy()
            elif '1Av' in valor:
                maq['1A-Motor Axial'] = extDados(l)
                maq_final[2] = maq.copy()
            elif '1He3' in valor:
                maq['1H-Motor Radial (env.)'] = extDados(l)
                maq_final[3] = maq.copy()
            elif '1Hv+' in valor:
                maq['1H-Motor Radial'] = extDados(l)
                maq_final[4] = maq.copy()
            elif '2Ve3+' in valor:
                maq['2H-Motor Radial (env.)'] = extDados(l)
                maq_final[5] = maq.copy()
            elif '2Vv' in valor:
                maq['2H-Motor Radial'] = extDados(l)
                maq_final[6] = maq.copy()
            elif 'Axial' in valor:
                maq['Carc.Axial-Central'] = extDados(l)
                maq_final[7] = maq.copy()
            elif 'Radial' in valor:
                maq['Ponto'] = extCabecalho(l)
                maq_final[0] = maq.copy()
                maq.clear()
                maq['Carc.Radial-Central'] = extDados(l)
                maq_final[8] = maq.copy()
                nome_arq = nomeArq(l)  # Captura nome do arquivo
            elif 'BASE01' in valor:
                maq['Base-01'] = extDados(l)
                maq_final[9] = maq.copy()
            elif 'BASE02' in valor:
                maq['Base-02'] = extDados(l)
                maq_final[10] = maq.copy()
            elif 'BASE03' in valor:
                maq['Base-03'] = extDados(l)
                maq_final[11] = maq.copy()
            elif 'BASE04' in valor:
                maq['Base-04'] = extDados(l)
                maq_final[12] = maq.copy()
except:
    print('  => [ ERRO ] - Extração dos pontos')
else:
    print('  => [ OK ] - Extração dos pontos')

sleep(0.8)
try: # Gravação dados novo sheet "final"
    final = book['tabela-final']
    lin_final = col_final = 1
    for ponto in maq_final:
        for k, v in ponto.items():
            final.cell(row=lin_final, column=col_final).value = k
            col_final += 1
            for medida in v:
                final.cell(row=lin_final, column=col_final).value = medida
                col_final += 1
        lin_final += 1
        col_final = 1
except:
    print('  => [ ERRO ] - Criação tabela-final')
else:
    print('  => [ OK ] - Criação tabela-final')

sleep(0.8)
try: # Salvar arquivo formato xlsx
    book.save(f'{nome_arq}.xlsx')
except:
    dataName = nomeArqErr()
    book.save(f'{dataName}.xlsx')
    print('  => [ ERRO ] - Criação novo arquivo nome estação".xlsx"')
    print('     => [ OK ] - Criação novo arquivo ".xlsx" c/ data atual')
else:
    print('  => [ OK ] - Criação novo arquivo ".xlsx"')

print('-' * 42)
input('<< Pressione qualquer tecla para sair >>')