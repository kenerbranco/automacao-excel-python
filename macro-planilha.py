"""
Ae3 - Motor/Axial (envelope)
Av+ - Motor/Axial
He3 - Motor/Radial (envelope)
Hv+ - Motor/Radial
Radial - Carcaça/Radial
Axial - Carcaça/Axial
LAe3+ - Lewa/Atuem/Axial(envelope)
LAv+ - Lewa/Atuem/Axial
LRe3+ - Lewa/Atuem/Radial(envelope)
LRv+ - Lewa/Atuem/Radial
"""
import openpyxl
from datetime import date
from time import sleep

maq = dict()
maq_final = list()

# Leitura arquivo .xlsx
book = openpyxl.load_workbook("index.xlsx")

# Atribuindo sheets em variáveis
index = book['index']
book.create_sheet('tabela-final')

# maximo de linhas e colunas
max_lin = index.max_row
max_col = index.max_column

def extCabecalho(l):  # Armazena dados cabeçalho
    lst = list()
    for la in range(l, l+5):
        cel = index.cell(row=la, column=15)
        val = cel.value
        lst.append(val)
    lst.append('Unidade')
    return lst


def extDados(l):
    lst = list()
    for la in range(l, l+5):  # Armazena cinco medições do ponto
        cel = index.cell(row=la, column=7)
        val = cel.value
        lst.append(val)
    cel_unidade = index.cell(row=l, column=9)  # Unidade medida do ponto
    cel_dif = index.cell(row=l, column=11)  # Diferença em porcentagem da última medição
    lst.append(cel_unidade.value)
    lst.append(cel_dif.value)
    return lst


def nomeArq(l):  # Armeza nome do equipamento para salvar planilha posteriormente
    cel_nome = index.cell(row=l, column=1)
    valor = cel_nome.value.split()
    data = date.today()
    nome = valor[0] + '-' + valor[1] + '-' + str(data)
    return nome


# Programa principal
print('-' * 42)
print('{:^42}'.format(' Automação Planilha v1.0 '))
print('{:^42}'.format(' Ventilação L1 - L3 '))
print('-' * 42)

sleep(0.8)
try: # Novo valor de max_lin para range de apenas dados necessários
    for l in range(1, max_lin+1): 
        for c in range(1, max_col):
            celula = index.cell(row=l, column=c)
            valor = str(celula.value).split()
            if 'Notas' in valor:
                max_lin = l
except:
    print('  => \033[0;31m[ ERRO ]\033[m - Filtro linhas necessárias')
else:
    print('  => \033[0;32m[ OK ]\033[m - Filtro linhas necessárias')

sleep(0.8)
try: # Extração dos pontos
    for l in range(1, max_lin+1):
        for c in range(1, max_col):
            maq.clear()
            celula = index.cell(row=l, column=c)
            valor = str(celula.value).split()
            if 'Ae3' in valor:
                maq['Motor/Axial(envelope)'] = extDados(l)
                maq_final.insert(1, maq.copy())
            elif 'Av+' in valor:
                maq['Motor/Axial'] = extDados(l)
                maq_final.insert(2, maq.copy())
            elif 'He3' in valor:
                maq['Motor/Radial(envelope)'] = extDados(l)
                maq_final.insert(3, maq.copy())
            elif 'Hv+' in valor:
                maq['Motor/Radial'] = extDados(l)
                maq_final.insert(4, maq.copy())
            elif 'Radial' in valor:
                maq['Ponto'] = extCabecalho(l)
                maq_final.insert(0, maq.copy())
                maq.clear()
                maq['Carcaça/Radial'] = extDados(l)
                maq_final.insert(5, maq.copy())
                nome_arq = nomeArq(l)  # Captura nome do arquivo
            elif 'Axial' in valor:
                maq['Carcaça/Axial'] = extDados(l)
                maq_final.insert(6, maq.copy())
            elif 'LAe3+' in valor:
                maq['Lewa/Atuem/Axial(envelope)'] = extDados(l)
                maq_final.insert(7, maq.copy())
            elif 'LAv+' in valor:
                maq['Lewa/Atuem/Axial'] = extDados(l)
                maq_final.insert(8, maq.copy())
            elif 'LRe3+' in valor:
                maq['Lewa/Atuem/Radial(envelope)'] = extDados(l)
                maq_final.insert(9, maq.copy())
            elif 'LRv+' in valor:
                maq['Leva/Atuem/Radial'] = extDados(l)
                maq_final.insert(10, maq.copy())
except:
    print('  => \033[0;31m[ ERRO ]\033[m - Extração dos pontos')
else:
    print('  => \033[0;32m[ OK ]\033[m - Extração dos pontos')

sleep(0.8)
try: # Gravação dados novo sheet "final"
    final = book['tabela-final']
    lin_final = col_final = 1
    for ponto in maq_final:
        for k, v in ponto.items():
            final.cell(row=lin_final, column=col_final).value = k
            col_final += 1
            for medida in v:
                final.cell(row=lin_final, column=col_final).value = medida
                col_final += 1
        lin_final += 1
        col_final = 1
    book.save(f'{nome_arq}.xlsx')
except:
    print('  => \033[0;31m[ ERRO ]\033[m - Criação novo arquivo ".xlsx"')
else:
    print('  => \033[0;32m[ OK ]\033[m - Criação novo arquivo ".xlsx"')
print('-' * 42)